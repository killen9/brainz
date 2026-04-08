#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
사내 CRM 시스템 - 인콜 고객사 관리
Flask + SQLite 기반 로컬 서버
"""

import os
import sqlite3
import json
import io
from datetime import datetime, date
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder='.')
CORS(app)

DB_PATH = os.path.join(os.path.dirname(__file__), 'crm.db')

# ─────────────────────────────────────────
# 데이터베이스 초기화
# ─────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    conn = get_db()
    cur = conn.cursor()

    # 고객사 테이블
    cur.execute('''
        CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            final_client       TEXT,          -- 최종 고객사
            inquiry_company    TEXT NOT NULL, -- 인콜 문의 회사
            inquiry_contact    TEXT,          -- 인콜 문의 담당자
            industry           TEXT,          -- 업종
            contract_status    TEXT DEFAULT '잠재',  -- 계약 상태 (잠재/상담중/견적/계약완료/보류/종결)
            inflow_route       TEXT,          -- 유입 경로
            callback_phone     TEXT,          -- 회신 연락처
            callback_email     TEXT,          -- 회신 이메일
            manager            TEXT,          -- 담당자 (사내)
            contract_info      TEXT,          -- 계약 정보 (JSON)
            memo               TEXT,          -- 메모
            created_at         TEXT DEFAULT (datetime('now','localtime')),
            updated_at         TEXT DEFAULT (datetime('now','localtime'))
        )
    ''')

    # 이력 테이블
    cur.execute('''
        CREATE TABLE IF NOT EXISTS histories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id   INTEGER NOT NULL,
            type        TEXT NOT NULL,  -- 통화/이메일/방문/미팅/기타
            title       TEXT NOT NULL,
            content     TEXT,
            contact     TEXT,           -- 연락한 담당자
            result      TEXT,           -- 결과
            created_by  TEXT,           -- 작성자
            created_at  TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE CASCADE
        )
    ''')

    # 일정 테이블
    cur.execute('''
        CREATE TABLE IF NOT EXISTS schedules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id   INTEGER,
            title       TEXT NOT NULL,
            description TEXT,
            start_dt    TEXT NOT NULL,  -- 시작 일시
            end_dt      TEXT,           -- 종료 일시
            type        TEXT DEFAULT '미팅',  -- 미팅/통화/방문/제안/기타
            status      TEXT DEFAULT '예정',  -- 예정/완료/취소
            manager     TEXT,           -- 담당자
            created_at  TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE SET NULL
        )
    ''')

    conn.commit()
    conn.close()
    print("[DB] 데이터베이스 초기화 완료")

def insert_sample_data():
    conn = get_db()
    cur = conn.cursor()

    # 이미 데이터가 있으면 스킵
    cur.execute("SELECT COUNT(*) FROM clients")
    if cur.fetchone()[0] > 0:
        conn.close()
        return

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    today = date.today().isoformat()

    clients = [
        ('삼성전자', '(주)테크솔루션', '김민준 차장', 'IT/제조', '상담중', '홈페이지', '010-1234-5678', 'kim@techsolution.co.kr', '이영희', '{"amount":"5000만원","period":"2024-06"}', '솔루션 도입 문의'),
        ('현대자동차', '오토파트너스', '박서연 팀장', '자동차/부품', '견적', '지인소개', '010-2345-6789', 'park@autopartners.co.kr', '최준혁', '{"amount":"3000만원","period":"2024-07"}', '시스템 개선 요청'),
        ('LG화학', '케미컬트레이드', '이지훈 과장', '화학/소재', '잠재', '광고', '010-3456-7890', 'lee@chemtrade.co.kr', '이영희', '{}', '초기 문의 단계'),
        ('SK하이닉스', '반도체파트너', '정수아 대리', '반도체', '계약완료', '전시회', '010-4567-8901', 'jung@semicond.co.kr', '최준혁', '{"amount":"1억원","period":"2024-01","end":"2024-12"}', '연간 계약 완료'),
        ('카카오', '디지털마케팅Co', '한예진 사원', 'IT/마케팅', '보류', '콜드콜', '010-5678-9012', 'han@digitalmkt.co.kr', '이영희', '{}', '예산 확보 후 재연락'),
        ('네이버', '클라우드파트너스', '강도현 부장', 'IT/클라우드', '상담중', '홈페이지', '010-6789-0123', 'kang@cloudp.co.kr', '최준혁', '{"amount":"2000만원"}', '클라우드 전환 검토'),
        ('롯데쇼핑', '리테일솔루션', '조민서 차장', '유통/리테일', '잠재', '세미나', '010-7890-1234', 'cho@retailsol.co.kr', '이영희', '{}', '디지털 전환 관심'),
        ('포스코', '스틸테크', '임채원 팀장', '철강/소재', '견적', '지인소개', '010-8901-2345', 'lim@steeltech.co.kr', '최준혁', '{"amount":"8000만원","period":"2024-08"}', '스마트팩토리 도입'),
    ]

    for c in clients:
        cur.execute('''INSERT INTO clients 
            (final_client,inquiry_company,inquiry_contact,industry,contract_status,
             inflow_route,callback_phone,callback_email,manager,contract_info,memo)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)''', c)

    client_ids = [cur.lastrowid - i for i in range(len(clients)-1, -1, -1)]

    histories = [
        (client_ids[0], '통화', '초기 상담 통화', 'ERP 솔루션 도입 관련 요구사항 청취', '김민준 차장', '추가 미팅 요청', '이영희'),
        (client_ids[0], '이메일', '제안서 발송', '맞춤형 솔루션 제안서 발송 완료', '김민준 차장', '검토 후 회신 예정', '이영희'),
        (client_ids[1], '미팅', '대면 미팅 진행', '요구사항 정리 및 견적 협의', '박서연 팀장', '견적서 요청', '최준혁'),
        (client_ids[3], '통화', '계약 최종 확인', '계약서 검토 완료 및 서명 예정', '정수아 대리', '계약 완료', '최준혁'),
        (client_ids[5], '이메일', '서비스 소개 자료 발송', '클라우드 마이그레이션 솔루션 소개', '강도현 부장', '관심 표명', '최준혁'),
    ]
    for h in histories:
        cur.execute('''INSERT INTO histories (client_id,type,title,content,contact,result,created_by)
            VALUES (?,?,?,?,?,?,?)''', h)

    # 일정 (오늘 기준 상대 날짜)
    schedules = [
        (client_ids[0], '삼성전자 2차 미팅', '솔루션 데모 및 기술 협의', f'{today}T14:00:00', f'{today}T16:00:00', '미팅', '예정', '이영희'),
        (client_ids[1], '현대자동차 견적 발표', '최종 견적안 프레젠테이션', f'{today}T10:00:00', f'{today}T12:00:00', '제안', '예정', '최준혁'),
        (client_ids[5], '네이버 전화 상담', '클라우드 전환 타임라인 협의', f'{today}T15:00:00', f'{today}T15:30:00', '통화', '예정', '최준혁'),
        (client_ids[7], '포스코 현장 방문', '스마트팩토리 현장 실사', '2026-04-12T09:00:00', '2026-04-12T18:00:00', '방문', '예정', '최준혁'),
        (client_ids[2], 'LG화학 초기 상담', '신규 문의 고객 첫 상담', '2026-04-10T11:00:00', '2026-04-10T12:00:00', '미팅', '예정', '이영희'),
        (client_ids[6], '롯데쇼핑 세미나 후속 연락', '세미나 참석 후 관심사항 확인', '2026-04-15T14:00:00', '2026-04-15T15:00:00', '통화', '예정', '이영희'),
        (client_ids[3], 'SK하이닉스 계약 완료', '연간 계약 서명 완료', '2026-03-28T14:00:00', '2026-03-28T15:00:00', '미팅', '완료', '최준혁'),
    ]
    for s in schedules:
        cur.execute('''INSERT INTO schedules (client_id,title,description,start_dt,end_dt,type,status,manager)
            VALUES (?,?,?,?,?,?,?,?)''', s)

    conn.commit()
    conn.close()
    print("[DB] 샘플 데이터 입력 완료")

# ─────────────────────────────────────────
# 정적 파일 서빙
# ─────────────────────────────────────────
@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/<path:filename>')
def static_files(filename):
    return send_from_directory('.', filename)

# ─────────────────────────────────────────
# 고객사 API
# ─────────────────────────────────────────
@app.route('/api/clients', methods=['GET'])
def get_clients():
    conn = get_db()
    cur = conn.cursor()

    search = request.args.get('search', '')
    status = request.args.get('status', '')
    manager = request.args.get('manager', '')
    industry = request.args.get('industry', '')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))

    where = []
    params = []

    if search:
        where.append('''(inquiry_company LIKE ? OR final_client LIKE ? OR 
                         inquiry_contact LIKE ? OR callback_phone LIKE ? OR manager LIKE ?)''')
        s = f'%{search}%'
        params.extend([s, s, s, s, s])
    if status:
        where.append('contract_status = ?')
        params.append(status)
    if manager:
        where.append('manager = ?')
        params.append(manager)
    if industry:
        where.append('industry LIKE ?')
        params.append(f'%{industry}%')

    where_str = ('WHERE ' + ' AND '.join(where)) if where else ''

    # 전체 수
    cur.execute(f'SELECT COUNT(*) FROM clients {where_str}', params)
    total = cur.fetchone()[0]

    # 페이지 조회
    offset = (page - 1) * per_page
    cur.execute(f'''SELECT * FROM clients {where_str} 
                    ORDER BY updated_at DESC 
                    LIMIT ? OFFSET ?''',
                params + [per_page, offset])
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()

    return jsonify({'total': total, 'page': page, 'per_page': per_page, 'data': rows})

@app.route('/api/clients/<int:cid>', methods=['GET'])
def get_client(cid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM clients WHERE id = ?', (cid,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return jsonify({'error': '고객사를 찾을 수 없습니다'}), 404
    return jsonify(dict(row))

@app.route('/api/clients', methods=['POST'])
def create_client():
    data = request.json
    conn = get_db()
    cur = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cur.execute('''INSERT INTO clients 
        (final_client,inquiry_company,inquiry_contact,industry,contract_status,
         inflow_route,callback_phone,callback_email,manager,contract_info,memo,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (data.get('final_client',''), data.get('inquiry_company',''), data.get('inquiry_contact',''),
         data.get('industry',''), data.get('contract_status','잠재'), data.get('inflow_route',''),
         data.get('callback_phone',''), data.get('callback_email',''), data.get('manager',''),
         json.dumps(data.get('contract_info', {}), ensure_ascii=False),
         data.get('memo',''), now, now))
    new_id = cur.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': new_id, 'message': '고객사가 등록되었습니다'}), 201

@app.route('/api/clients/<int:cid>', methods=['PUT'])
def update_client(cid):
    data = request.json
    conn = get_db()
    cur = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cur.execute('''UPDATE clients SET
        final_client=?, inquiry_company=?, inquiry_contact=?, industry=?,
        contract_status=?, inflow_route=?, callback_phone=?, callback_email=?,
        manager=?, contract_info=?, memo=?, updated_at=?
        WHERE id=?''',
        (data.get('final_client',''), data.get('inquiry_company',''), data.get('inquiry_contact',''),
         data.get('industry',''), data.get('contract_status','잠재'), data.get('inflow_route',''),
         data.get('callback_phone',''), data.get('callback_email',''), data.get('manager',''),
         json.dumps(data.get('contract_info', {}), ensure_ascii=False),
         data.get('memo',''), now, cid))
    conn.commit()
    conn.close()
    return jsonify({'message': '고객사 정보가 수정되었습니다'})

@app.route('/api/clients/<int:cid>', methods=['DELETE'])
def delete_client(cid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('DELETE FROM clients WHERE id = ?', (cid,))
    conn.commit()
    conn.close()
    return jsonify({'message': '고객사가 삭제되었습니다'})

# ─────────────────────────────────────────
# 이력 API
# ─────────────────────────────────────────
@app.route('/api/clients/<int:cid>/histories', methods=['GET'])
def get_histories(cid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM histories WHERE client_id=? ORDER BY created_at DESC', (cid,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify(rows)

@app.route('/api/histories', methods=['GET'])
def get_all_histories():
    conn = get_db()
    cur = conn.cursor()
    cur.execute('''SELECT h.*, c.inquiry_company, c.final_client 
                   FROM histories h LEFT JOIN clients c ON h.client_id = c.id
                   ORDER BY h.created_at DESC LIMIT 100''')
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify(rows)

@app.route('/api/histories', methods=['POST'])
def create_history():
    data = request.json
    conn = get_db()
    cur = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cur.execute('''INSERT INTO histories (client_id,type,title,content,contact,result,created_by,created_at)
        VALUES (?,?,?,?,?,?,?,?)''',
        (data.get('client_id'), data.get('type','기타'), data.get('title',''),
         data.get('content',''), data.get('contact',''), data.get('result',''),
         data.get('created_by',''), now))
    new_id = cur.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': new_id, 'message': '이력이 등록되었습니다'}), 201

@app.route('/api/histories/<int:hid>', methods=['PUT'])
def update_history(hid):
    data = request.json
    conn = get_db()
    cur = conn.cursor()
    cur.execute('''UPDATE histories SET type=?,title=?,content=?,contact=?,result=?,created_by=?
        WHERE id=?''',
        (data.get('type','기타'), data.get('title',''), data.get('content',''),
         data.get('contact',''), data.get('result',''), data.get('created_by',''), hid))
    conn.commit()
    conn.close()
    return jsonify({'message': '이력이 수정되었습니다'})

@app.route('/api/histories/<int:hid>', methods=['DELETE'])
def delete_history(hid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('DELETE FROM histories WHERE id = ?', (hid,))
    conn.commit()
    conn.close()
    return jsonify({'message': '이력이 삭제되었습니다'})

# ─────────────────────────────────────────
# 일정 API
# ─────────────────────────────────────────
@app.route('/api/schedules', methods=['GET'])
def get_schedules():
    conn = get_db()
    cur = conn.cursor()
    status = request.args.get('status', '')
    manager = request.args.get('manager', '')
    from_dt = request.args.get('from', '')
    to_dt = request.args.get('to', '')

    where = []
    params = []
    if status:
        where.append('s.status = ?')
        params.append(status)
    if manager:
        where.append('s.manager = ?')
        params.append(manager)
    if from_dt:
        where.append('s.start_dt >= ?')
        params.append(from_dt)
    if to_dt:
        where.append('s.start_dt <= ?')
        params.append(to_dt + 'T23:59:59')

    where_str = ('WHERE ' + ' AND '.join(where)) if where else ''
    cur.execute(f'''SELECT s.*, c.inquiry_company, c.final_client
                    FROM schedules s LEFT JOIN clients c ON s.client_id = c.id
                    {where_str} ORDER BY s.start_dt ASC''', params)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify(rows)

@app.route('/api/schedules', methods=['POST'])
def create_schedule():
    data = request.json
    conn = get_db()
    cur = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cur.execute('''INSERT INTO schedules (client_id,title,description,start_dt,end_dt,type,status,manager,created_at)
        VALUES (?,?,?,?,?,?,?,?,?)''',
        (data.get('client_id'), data.get('title',''), data.get('description',''),
         data.get('start_dt',''), data.get('end_dt',''), data.get('type','미팅'),
         data.get('status','예정'), data.get('manager',''), now))
    new_id = cur.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': new_id, 'message': '일정이 등록되었습니다'}), 201

@app.route('/api/schedules/<int:sid>', methods=['PUT'])
def update_schedule(sid):
    data = request.json
    conn = get_db()
    cur = conn.cursor()
    cur.execute('''UPDATE schedules SET client_id=?,title=?,description=?,start_dt=?,end_dt=?,
        type=?,status=?,manager=? WHERE id=?''',
        (data.get('client_id'), data.get('title',''), data.get('description',''),
         data.get('start_dt',''), data.get('end_dt',''), data.get('type','미팅'),
         data.get('status','예정'), data.get('manager',''), sid))
    conn.commit()
    conn.close()
    return jsonify({'message': '일정이 수정되었습니다'})

@app.route('/api/schedules/<int:sid>', methods=['DELETE'])
def delete_schedule(sid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('DELETE FROM schedules WHERE id = ?', (sid,))
    conn.commit()
    conn.close()
    return jsonify({'message': '일정이 삭제되었습니다'})

# ─────────────────────────────────────────
# 대시보드 통계 API
# ─────────────────────────────────────────
@app.route('/api/dashboard', methods=['GET'])
def get_dashboard():
    conn = get_db()
    cur = conn.cursor()
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    today_str = date.today().isoformat()
    week_end = datetime.now().strftime('%Y-%m-%d')

    # 상태별 고객사 수
    cur.execute('''SELECT contract_status, COUNT(*) as cnt FROM clients GROUP BY contract_status''')
    status_stats = {r['contract_status']: r['cnt'] for r in cur.fetchall()}

    # 전체 고객사 수
    cur.execute('SELECT COUNT(*) FROM clients')
    total_clients = cur.fetchone()[0]

    # 이번달 신규 고객사
    cur.execute("""SELECT COUNT(*) FROM clients 
                   WHERE strftime('%Y-%m', created_at) = strftime('%Y-%m', 'now', 'localtime')""")
    new_this_month = cur.fetchone()[0]

    # 향후 30일 일정
    cur.execute('''SELECT s.*, c.inquiry_company, c.final_client
                   FROM schedules s LEFT JOIN clients c ON s.client_id = c.id
                   WHERE s.start_dt >= ? AND s.start_dt <= date(?, '+30 days') AND s.status = '예정'
                   ORDER BY s.start_dt ASC LIMIT 20''', (today_str, today_str))
    upcoming = [dict(r) for r in cur.fetchall()]

    # 진행중 고객사 (상담중 + 견적)
    cur.execute('''SELECT c.*, 
                   (SELECT COUNT(*) FROM histories h WHERE h.client_id = c.id) as history_count,
                   (SELECT MAX(h2.created_at) FROM histories h2 WHERE h2.client_id = c.id) as last_contact
                   FROM clients c 
                   WHERE c.contract_status IN ('상담중','견적')
                   ORDER BY c.updated_at DESC''')
    in_progress = [dict(r) for r in cur.fetchall()]

    # 최근 이력 5건
    cur.execute('''SELECT h.*, c.inquiry_company FROM histories h 
                   LEFT JOIN clients c ON h.client_id = c.id
                   ORDER BY h.created_at DESC LIMIT 5''')
    recent_histories = [dict(r) for r in cur.fetchall()]

    # 담당자별 통계
    cur.execute('''SELECT manager, COUNT(*) as cnt FROM clients 
                   WHERE manager != '' GROUP BY manager ORDER BY cnt DESC''')
    manager_stats = [dict(r) for r in cur.fetchall()]

    conn.close()
    return jsonify({
        'status_stats': status_stats,
        'total_clients': total_clients,
        'new_this_month': new_this_month,
        'upcoming_schedules': upcoming,
        'in_progress_clients': in_progress,
        'recent_histories': recent_histories,
        'manager_stats': manager_stats
    })

# ─────────────────────────────────────────
# 엑셀 Export API
# ─────────────────────────────────────────
@app.route('/api/export/clients', methods=['GET'])
def export_clients():
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM clients ORDER BY id')
    rows = cur.fetchall()

    # 이력도 함께 export
    include_history = request.args.get('include_history', 'false') == 'true'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '고객사 목록'

    # 스타일 정의
    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    status_colors = {
        '잠재': 'EFF6FF', '상담중': 'FEF3C7', '견적': 'FFF7ED',
        '계약완료': 'D1FAE5', '보류': 'FEE2E2', '종결': 'F3F4F6'
    }

    headers = [
        'ID', '최종 고객사', '인콜 문의 회사', '인콜 문의 담당자', '업종',
        '계약 상태', '유입 경로', '회신 연락처', '회신 이메일', '담당자(사내)',
        '계약 금액', '계약 기간', '메모', '등록일', '최종 수정일'
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    col_widths = [6, 18, 20, 15, 12, 10, 12, 15, 25, 10, 12, 12, 30, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in rows:
        d = dict(row)
        try:
            ci = json.loads(d.get('contract_info') or '{}')
        except:
            ci = {}
        row_data = [
            d['id'], d.get('final_client',''), d.get('inquiry_company',''),
            d.get('inquiry_contact',''), d.get('industry',''), d.get('contract_status',''),
            d.get('inflow_route',''), d.get('callback_phone',''), d.get('callback_email',''),
            d.get('manager',''), ci.get('amount',''), ci.get('period',''),
            d.get('memo',''), d.get('created_at',''), d.get('updated_at','')
        ]
        ws.append(row_data)
        status = d.get('contract_status','')
        if status in status_colors:
            fill = PatternFill(start_color=status_colors[status], end_color=status_colors[status], fill_type='solid')
            for cell in ws[ws.max_row]:
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # 이력 시트
    if include_history:
        cur.execute('''SELECT h.*, c.inquiry_company FROM histories h 
                       LEFT JOIN clients c ON h.client_id = c.id ORDER BY h.client_id, h.created_at''')
        hist_rows = cur.fetchall()

        ws2 = wb.create_sheet('이력 목록')
        h_headers = ['ID', '고객사', '유형', '제목', '내용', '연락 담당자', '결과', '작성자', '일시']
        ws2.append(h_headers)
        for cell in ws2[1]:
            cell.fill = PatternFill(start_color='065F46', end_color='065F46', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for hr in hist_rows:
            hd = dict(hr)
            ws2.append([hd['id'], hd.get('inquiry_company',''), hd.get('type',''),
                         hd.get('title',''), hd.get('content',''), hd.get('contact',''),
                         hd.get('result',''), hd.get('created_by',''), hd.get('created_at','')])

        h_widths = [6, 20, 8, 25, 40, 15, 20, 10, 18]
        for i, w in enumerate(h_widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

    conn.close()
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"고객사목록_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

# ─────────────────────────────────────────
# 엑셀 Import API
# ─────────────────────────────────────────
@app.route('/api/import/clients', methods=['POST'])
def import_clients():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '엑셀 파일(.xlsx, .xls)만 업로드 가능합니다'}), 400

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        return jsonify({'error': f'파일 읽기 오류: {str(e)}'}), 400

    conn = get_db()
    cur = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    success, errors = 0, []

    for i, row in enumerate(rows, 2):
        if not row or not row[0]:
            continue
        try:
            # 컬럼: 최종고객사/문의회사/담당자/업종/상태/유입경로/연락처/이메일/담당자/금액/기간/메모
            final_client     = str(row[0]).strip() if row[0] else ''
            inquiry_company  = str(row[1]).strip() if row[1] else ''
            inquiry_contact  = str(row[2]).strip() if row[2] else ''
            industry         = str(row[3]).strip() if row[3] else ''
            status           = str(row[4]).strip() if row[4] else '잠재'
            inflow_route     = str(row[5]).strip() if row[5] else ''
            callback_phone   = str(row[6]).strip() if row[6] else ''
            callback_email   = str(row[7]).strip() if row[7] else ''
            manager          = str(row[8]).strip() if row[8] else ''
            amount           = str(row[9]).strip() if row[9] else ''
            period           = str(row[10]).strip() if len(row) > 10 and row[10] else ''
            memo             = str(row[11]).strip() if len(row) > 11 and row[11] else ''

            if not inquiry_company:
                errors.append(f'행 {i}: 인콜 문의 회사가 없습니다')
                continue

            contract_info = json.dumps({'amount': amount, 'period': period}, ensure_ascii=False)
            valid_statuses = ['잠재','상담중','견적','계약완료','보류','종결']
            if status not in valid_statuses:
                status = '잠재'

            cur.execute('''INSERT INTO clients 
                (final_client,inquiry_company,inquiry_contact,industry,contract_status,
                 inflow_route,callback_phone,callback_email,manager,contract_info,memo,created_at,updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (final_client, inquiry_company, inquiry_contact, industry, status,
                 inflow_route, callback_phone, callback_email, manager, contract_info, memo, now, now))
            success += 1
        except Exception as e:
            errors.append(f'행 {i}: {str(e)}')

    conn.commit()
    conn.close()
    return jsonify({'success': success, 'errors': errors, 'message': f'{success}건 등록 완료'})

# ─────────────────────────────────────────
# 엑셀 템플릿 다운로드
# ─────────────────────────────────────────
@app.route('/api/export/template', methods=['GET'])
def export_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '고객사 입력'

    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)

    headers = [
        '최종 고객사', '인콜 문의 회사*', '인콜 문의 담당자', '업종',
        '계약 상태', '유입 경로', '회신 연락처', '회신 이메일',
        '사내 담당자', '계약 금액', '계약 기간', '메모'
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 예시 데이터
    examples = [
        ['삼성전자', '(주)테크파트너', '홍길동 대리', 'IT/제조', '잠재', '홈페이지',
         '010-0000-0000', 'hong@example.com', '이담당', '1000만원', '2024-12', ''],
    ]
    for ex in examples:
        ws.append(ex)

    widths = [18, 20, 15, 12, 10, 12, 15, 25, 10, 12, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 안내 시트
    ws2 = wb.create_sheet('입력 안내')
    ws2.append(['항목', '설명', '필수여부', '예시값'])
    guide = [
        ['최종 고객사', '실제 최종 납품/서비스 대상 고객사', '선택', '삼성전자'],
        ['인콜 문의 회사', '전화/문의한 회사명', '필수', '(주)테크파트너'],
        ['인콜 문의 담당자', '문의한 담당자 이름 및 직책', '선택', '홍길동 대리'],
        ['업종', '고객사 업종', '선택', 'IT/제조, 자동차, 화학 등'],
        ['계약 상태', '잠재/상담중/견적/계약완료/보류/종결', '선택', '잠재'],
        ['유입 경로', '인콜 유입 경로', '선택', '홈페이지, 지인소개, 광고, 전시회 등'],
        ['회신 연락처', '연락 가능한 전화번호', '선택', '010-0000-0000'],
        ['회신 이메일', '연락 가능한 이메일', '선택', 'hong@example.com'],
        ['사내 담당자', '담당 영업/CS 직원', '선택', '이영희'],
        ['계약 금액', '예상 또는 확정 계약 금액', '선택', '1000만원'],
        ['계약 기간', '계약 기간 또는 예상 기간', '선택', '2024-12'],
        ['메모', '기타 메모사항', '선택', ''],
    ]
    for g in guide:
        ws2.append(g)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='고객사_입력양식.xlsx')

# ─────────────────────────────────────────
# 공통 유틸
# ─────────────────────────────────────────
@app.route('/api/managers', methods=['GET'])
def get_managers():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT manager FROM clients WHERE manager != '' ORDER BY manager")
    managers = [r['manager'] for r in cur.fetchall()]
    conn.close()
    return jsonify(managers)

if __name__ == '__main__':
    init_db()
    insert_sample_data()
    print("=" * 50)
    print(" CRM 서버 시작: http://localhost:5000")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5000, debug=False)
